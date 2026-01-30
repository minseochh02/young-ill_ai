# -*- coding: utf-8 -*-
"""
영일오엔씨 일보현황 엑셀 파일 파싱 스크립트
- 여러 날짜의 엑셀 파일을 읽어 JSON으로 변환
- 일별 데이터 및 누계 데이터 생성
- 마스터 DB 엑셀 파일 생성
"""

import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE_PATH = r'c:\Users\JS\Desktop\0101\ILBO'
OUTPUT_PATH = r'c:\Users\JS\Desktop\0101\daily-report-dashboard\src\data'
MASTER_DB_PATH = r'c:\Users\JS\Desktop\0101\ILBO'
PUBLIC_PATH = r'c:\Users\JS\Desktop\0101\daily-report-dashboard\public'

# 파일 목록 (날짜순)
FILES = [
    ('2025-12-02', '영일오엔씨 일보현황 2025년 - 1202.xlsx'),
    ('2025-12-03', '영일오엔씨 일보현황 2025년 - 1203.xlsx'),
    ('2025-12-04', '영일오엔씨 일보현황 2025년 - 1204.xlsx'),
    ('2025-12-05', '영일오엔씨 일보현황 2025년 - 1205.xlsx'),
    ('2025-12-06', '영일오엔씨 일보현황 2025년 - 1206.xlsx'),
    ('2025-12-07', '영일오엔씨 일보현황 2025년 - 1207.xlsx'),
]

def safe_float(val):
    """안전하게 float로 변환"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def parse_sales_data(df_ilbo):
    """매출현황 파싱 (일보 시트 Row 5-12)"""
    sales = []
    branches = ['서울,화성 IL', '창원', '화성auto(남부)', '화성auto(중부)', '인천(서부)', '남양주(동부)', '제주', '부산']

    for i, branch in enumerate(branches):
        row_idx = 5 + i  # Row 5부터 시작
        if row_idx < len(df_ilbo):
            row = df_ilbo.iloc[row_idx]
            sales.append({
                'branch': branch,
                'total_sales': safe_float(row.iloc[1]),
                'mobil_sell_out': safe_float(row.iloc[2]),
                'mobil_sell_out_total_l': safe_float(row.iloc[3]),
                'mobil_sell_out_flagship_l': safe_float(row.iloc[4]),
                'mobil_sell_in_total_l': safe_float(row.iloc[5]),
                'mobil_sell_in_flagship_l': safe_float(row.iloc[7]),
            })
    return sales

def parse_collections_data(df_ilbo):
    """외상매출금 현황 파싱 (일보 시트 Row 19-26)"""
    collections = []
    branches = ['화성 IL', '창원', '화성auto(남부)', '화성auto(중부)', '인천(서부)', '남양주(동부)', '제주', '부산']

    for i, branch in enumerate(branches):
        row_idx = 19 + i  # Row 19부터 시작
        if row_idx < len(df_ilbo):
            row = df_ilbo.iloc[row_idx]
            collections.append({
                'branch': branch,
                'total_collection': safe_float(row.iloc[1]),
                'cash': safe_float(row.iloc[2]),
                'bill': safe_float(row.iloc[3]),
                'card': safe_float(row.iloc[4]),
                'etc1': safe_float(row.iloc[5]),
                'etc2': safe_float(row.iloc[6]),
                'prev_balance': safe_float(row.iloc[12]),
                'current_sales': safe_float(row.iloc[13]),
                'ar_balance': safe_float(row.iloc[14]),
            })
    return collections

def parse_funds_data(df_ilbo):
    """자금 현황 파싱 (일보 시트 Row 32-35)"""
    funds = []
    fund_types = ['전잔', '당입', '지출', '현잔']

    for i, fund_type in enumerate(fund_types):
        row_idx = 32 + i  # Row 32부터 시작
        if row_idx < len(df_ilbo):
            row = df_ilbo.iloc[row_idx]
            funds.append({
                'type': fund_type,
                'general_deposit': safe_float(row.iloc[1]),
                'electronic_bill': safe_float(row.iloc[2]),
                'loans': safe_float(row.iloc[3]),
                'receivable_bill': safe_float(row.iloc[4]),
                'savings_insurance': safe_float(row.iloc[5]),
                'cma': safe_float(row.iloc[6]),
                'foreign_usd': safe_float(row.iloc[7]),
                'foreign_eur': safe_float(row.iloc[8]),
                'foreign_jpy': safe_float(row.iloc[9]),
                'foreign_usd_normal': safe_float(row.iloc[10]),
                'credit_limit': safe_float(row.iloc[11]),
                'short_term_loan': safe_float(row.iloc[12]),
                'long_term_loan': safe_float(row.iloc[13]),
                'retirement_pension': safe_float(row.iloc[14]),
            })
    return funds

def parse_major_expenses(df_ilbo):
    """주요비용 지출현황 파싱 (일보 시트 Row 43-52)"""
    expenses = []

    for row_idx in range(43, 53):  # Row 43부터 시작
        if row_idx < len(df_ilbo):
            row = df_ilbo.iloc[row_idx]
            client = row.iloc[4] if pd.notna(row.iloc[4]) else None
            amount = row.iloc[5] if pd.notna(row.iloc[5]) else None
            desc = row.iloc[6] if pd.notna(row.iloc[6]) else None

            if client and pd.notna(amount):
                expenses.append({
                    'client': str(client),
                    'amount': safe_float(amount),
                    'description': str(desc) if pd.notna(desc) else ''
                })
    return expenses

def parse_mobil_payments(df_ilbo):
    """모빌결제내역 파싱 (일보 시트 Row 44-51)"""
    payments = []
    branches_map = {
        '화성 IL': 44,
        '창원 IL': 45,
        '화성 AUTO (중부)': 46,
        '남부지사': 47,
        '인천(서부)': 48,
        '남양주(동부)': 49,
        '제주': 50,
        '부산': 51,
    }

    for branch, row_idx in branches_map.items():
        if row_idx < len(df_ilbo):
            row = df_ilbo.iloc[row_idx]
            payments.append({
                'branch': branch,
                'il': safe_float(row.iloc[10]),
                'auto': safe_float(row.iloc[11]),
                'mbk': safe_float(row.iloc[12]),
                'total': safe_float(row.iloc[13]),
            })
    return payments

def parse_excel_file(file_path, date_str):
    """엑셀 파일 하나를 파싱하여 데이터 반환"""
    try:
        df_ilbo = pd.read_excel(file_path, sheet_name='일보', header=None)

        data = {
            'date': date_str,
            'sales': parse_sales_data(df_ilbo),
            'collections': parse_collections_data(df_ilbo),
            'funds': parse_funds_data(df_ilbo),
            'major_expenses': parse_major_expenses(df_ilbo),
            'mobil_payments': parse_mobil_payments(df_ilbo),
        }

        # 합계 계산
        data['totals'] = {
            'total_sales': sum(s['total_sales'] for s in data['sales']),
            'total_mobil_sell_out': sum(s['mobil_sell_out'] for s in data['sales']),
            'total_collection': sum(c['total_collection'] for c in data['collections']),
            'total_ar_balance': sum(c['ar_balance'] for c in data['collections']),
        }

        return data
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        return None

def calculate_cumulative(daily_data_list):
    """누계 데이터 계산"""
    cumulative = {
        'sales': {},
        'collections': {},
        'funds': {
            'total_deposit': 0,  # 총 입금 누계
            'total_withdraw': 0,  # 총 출금 누계
        },
        'totals': {
            'total_sales': 0,
            'total_mobil_sell_out': 0,
            'total_collection': 0,
            'total_ar_balance': 0,
            'total_funds_in': 0,
        }
    }

    for daily in daily_data_list:
        # 매출 누계
        for sale in daily['sales']:
            branch = sale['branch']
            if branch not in cumulative['sales']:
                cumulative['sales'][branch] = {
                    'branch': branch,
                    'total_sales': 0,
                    'mobil_sell_out': 0,
                    'mobil_sell_out_total_l': 0,
                    'mobil_sell_out_flagship_l': 0,
                    'mobil_sell_in_total_l': 0,
                    'mobil_sell_in_flagship_l': 0,
                }
            cumulative['sales'][branch]['total_sales'] += sale['total_sales']
            cumulative['sales'][branch]['mobil_sell_out'] += sale['mobil_sell_out']
            cumulative['sales'][branch]['mobil_sell_out_total_l'] += sale['mobil_sell_out_total_l']
            cumulative['sales'][branch]['mobil_sell_out_flagship_l'] += sale['mobil_sell_out_flagship_l']
            cumulative['sales'][branch]['mobil_sell_in_total_l'] += sale['mobil_sell_in_total_l']
            cumulative['sales'][branch]['mobil_sell_in_flagship_l'] += sale['mobil_sell_in_flagship_l']

        # 수금 누계 (ar_balance는 마지막 값 사용)
        for coll in daily['collections']:
            branch = coll['branch']
            if branch not in cumulative['collections']:
                cumulative['collections'][branch] = {
                    'branch': branch,
                    'total_collection': 0,
                    'cash': 0,
                    'bill': 0,
                    'card': 0,
                    'ar_balance': 0,  # 마지막 날짜의 잔액
                }
            cumulative['collections'][branch]['total_collection'] += coll['total_collection']
            cumulative['collections'][branch]['cash'] += coll['cash']
            cumulative['collections'][branch]['bill'] += coll['bill']
            cumulative['collections'][branch]['card'] += coll['card']
            # ar_balance는 마지막 날짜 값으로 덮어씀 (누적이 아닌 현재 잔액)
            cumulative['collections'][branch]['ar_balance'] = coll['ar_balance']

        # 자금 누계
        for fund in daily['funds']:
            if fund['type'] == '당입':
                cumulative['funds']['total_deposit'] += fund['general_deposit']
            elif fund['type'] == '지출':
                cumulative['funds']['total_withdraw'] += fund['general_deposit']

        # 총계 누계
        cumulative['totals']['total_sales'] += daily['totals']['total_sales']
        cumulative['totals']['total_mobil_sell_out'] += daily['totals']['total_mobil_sell_out']
        cumulative['totals']['total_collection'] += daily['totals']['total_collection']

    # 마지막 날짜 기준 미수잔액 합계
    cumulative['totals']['total_ar_balance'] = sum(
        c['ar_balance'] for c in cumulative['collections'].values()
    )
    # 자금 입금 누계
    cumulative['totals']['total_funds_in'] = cumulative['funds']['total_deposit']

    # dict를 list로 변환
    cumulative['sales'] = list(cumulative['sales'].values())
    cumulative['collections'] = list(cumulative['collections'].values())

    return cumulative

def create_master_db_excel(all_data, output_path):
    """마스터 DB 엑셀 파일 생성"""
    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. 매출DB 시트
    ws_sales = wb.active
    ws_sales.title = "매출DB"
    sales_headers = ['날짜', '년', '월', '일', '사업소', '총매출액', '모빌Sell-out금액',
                     'Sell-out_Total(L)', 'Sell-out_Flagship(L)',
                     'Sell-in_Total(L)', 'Sell-in_Flagship(L)']
    ws_sales.append(sales_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        for sale in daily['sales']:
            ws_sales.append([
                date, int(year), int(month), int(day),
                sale['branch'],
                sale['total_sales'],
                sale['mobil_sell_out'],
                sale['mobil_sell_out_total_l'],
                sale['mobil_sell_out_flagship_l'],
                sale['mobil_sell_in_total_l'],
                sale['mobil_sell_in_flagship_l'],
            ])

    # 헤더 스타일 적용
    for cell in ws_sales[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 2. 수금DB 시트
    ws_coll = wb.create_sheet("수금DB")
    coll_headers = ['날짜', '년', '월', '일', '사업소', '총수금액', '현금', '어음', '카드', '전월잔액', '당월매출', '미수잔액']
    ws_coll.append(coll_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        for coll in daily['collections']:
            ws_coll.append([
                date, int(year), int(month), int(day),
                coll['branch'],
                coll['total_collection'],
                coll['cash'],
                coll['bill'],
                coll['card'],
                coll['prev_balance'],
                coll['current_sales'],
                coll['ar_balance'],
            ])

    for cell in ws_coll[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 3. 자금DB 시트
    ws_funds = wb.create_sheet("자금DB")
    funds_headers = ['날짜', '년', '월', '일', '구분', '보통예금', '전자어음', '외담대',
                     '받을어음', '적금+보험', 'CMA', '외화USD', '외화EUR', '외화JPY',
                     '외화보통USD', '한도대출', '단기차입금', '장기차입금', '퇴직연금']
    ws_funds.append(funds_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        for fund in daily['funds']:
            ws_funds.append([
                date, int(year), int(month), int(day),
                fund['type'],
                fund['general_deposit'],
                fund['electronic_bill'],
                fund['loans'],
                fund['receivable_bill'],
                fund['savings_insurance'],
                fund['cma'],
                fund['foreign_usd'],
                fund['foreign_eur'],
                fund['foreign_jpy'],
                fund['foreign_usd_normal'],
                fund['credit_limit'],
                fund['short_term_loan'],
                fund['long_term_loan'],
                fund['retirement_pension'],
            ])

    for cell in ws_funds[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 4. 주요비용DB 시트
    ws_exp = wb.create_sheet("주요비용DB")
    exp_headers = ['날짜', '년', '월', '일', '거래처', '금액', '적요']
    ws_exp.append(exp_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        for exp in daily['major_expenses']:
            ws_exp.append([
                date, int(year), int(month), int(day),
                exp['client'],
                exp['amount'],
                exp['description'],
            ])

    for cell in ws_exp[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 5. 모빌결제DB 시트
    ws_mobil = wb.create_sheet("모빌결제DB")
    mobil_headers = ['날짜', '년', '월', '일', '사업소', 'IL', 'AUTO', 'MBK', '합계']
    ws_mobil.append(mobil_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        for pay in daily['mobil_payments']:
            ws_mobil.append([
                date, int(year), int(month), int(day),
                pay['branch'],
                pay['il'],
                pay['auto'],
                pay['mbk'],
                pay['total'],
            ])

    for cell in ws_mobil[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 6. 일별합계DB 시트
    ws_daily = wb.create_sheet("일별합계DB")
    daily_headers = ['날짜', '년', '월', '일', '총매출액', '총Sell-out', '총수금액', '총미수잔액']
    ws_daily.append(daily_headers)

    for daily in all_data:
        date = daily['date']
        year, month, day = date.split('-')
        ws_daily.append([
            date, int(year), int(month), int(day),
            daily['totals']['total_sales'],
            daily['totals']['total_mobil_sell_out'],
            daily['totals']['total_collection'],
            daily['totals']['total_ar_balance'],
        ])

    for cell in ws_daily[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 컬럼 너비 조정
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"마스터 DB 엑셀 저장 완료: {output_path}")

def main():
    print("=" * 60)
    print("영일오엔씨 일보현황 파싱 시작")
    print("=" * 60)

    all_data = []

    for date_str, filename in FILES:
        file_path = os.path.join(BASE_PATH, filename)
        print(f"\n처리 중: {filename}")

        data = parse_excel_file(file_path, date_str)
        if data:
            all_data.append(data)
            print(f"  - 매출 데이터: {len(data['sales'])}건")
            print(f"  - 수금 데이터: {len(data['collections'])}건")
            print(f"  - 자금 데이터: {len(data['funds'])}건")
            print(f"  - 주요비용: {len(data['major_expenses'])}건")
            print(f"  - 총매출: {data['totals']['total_sales']:,.0f}원")

    # 누계 계산
    cumulative = calculate_cumulative(all_data)
    print(f"\n=== 누계 합산 결과 ===")
    print(f"총 매출 누계: {cumulative['totals']['total_sales']:,.0f}원")
    print(f"총 수금 누계: {cumulative['totals']['total_collection']:,.0f}원")

    # JSON 파일 저장
    os.makedirs(OUTPUT_PATH, exist_ok=True)

    # 일별 데이터 저장
    daily_output = {
        'generated_at': datetime.now().isoformat(),
        'dates': [d['date'] for d in all_data],
        'daily_data': all_data,
        'cumulative': cumulative,
    }

    json_path = os.path.join(OUTPUT_PATH, 'daily_reports.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(daily_output, f, ensure_ascii=False, indent=2)
    print(f"\nJSON 저장 완료: {json_path}")

    # 마스터 DB 엑셀 생성
    excel_path = os.path.join(MASTER_DB_PATH, '영일오엔씨_마스터DB.xlsx')
    create_master_db_excel(all_data, excel_path)

    # public 폴더에도 복사 (웹 다운로드용)
    import shutil
    os.makedirs(PUBLIC_PATH, exist_ok=True)
    public_excel_path = os.path.join(PUBLIC_PATH, '영일오엔씨_마스터DB.xlsx')
    shutil.copy(excel_path, public_excel_path)
    print(f"다운로드용 복사 완료: {public_excel_path}")

    print("\n" + "=" * 60)
    print("처리 완료!")
    print("=" * 60)

if __name__ == '__main__':
    main()
